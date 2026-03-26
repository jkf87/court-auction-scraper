#!/usr/bin/env python3
"""
Court Auction Excel Generator
Usage: python3 create_auction_excel.py <input_json> <output_xlsx> [--region REGION] [--usage USAGE]

input_json: JSON file — accepts either:
  - A flat array of item objects: [{...}, {...}, ...]
  - An object with "items" key: {"count": N, "items": [{...}, ...]}
  - An array of page results: [{"count": N, "items": [...]}, ...]
output_xlsx: Output Excel file path
--region: Filter by region keyword (e.g., "대전광역시")
--usage: Filter by usage keyword (e.g., "아파트")
"""
import json, argparse

REGION_MAP = [
    ('서울특별시', '서울'), ('서울', '서울'),
    ('부산광역시', '부산'), ('부산', '부산'),
    ('대구광역시', '대구'), ('대구', '대구'),
    ('인천광역시', '인천'), ('인천', '인천'),
    ('광주광역시', '광주'),
    ('대전광역시', '대전'), ('대전', '대전'),
    ('울산광역시', '울산'),
    ('세종특별자치시', '세종'),
    ('경기도', '경기'),
    ('강원특별자치도', '강원'), ('강원도', '강원'),
    ('충청북도', '충북'), ('충청남도', '충남'),
    ('전북특별자치도', '전북'), ('전라북도', '전북'),
    ('전라남도', '전남'),
    ('경상북도', '경북'), ('경상남도', '경남'),
    ('제주특별자치도', '제주'),
]

def get_region(addr):
    for keyword, label in REGION_MAP:
        if keyword in addr:
            return label
    return '기타'

def normalize_input(raw):
    """Accept multiple JSON formats and return a flat list of items."""
    if isinstance(raw, list):
        if len(raw) > 0 and isinstance(raw[0], dict) and 'items' in raw[0]:
            items = []
            for page in raw:
                items.extend(page.get('items', []))
            return items
        return raw
    if isinstance(raw, dict) and 'items' in raw:
        return raw['items']
    raise ValueError("Unrecognized JSON format")

def create_excel(items, output_path, region_filter=None, usage_filter=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_font = Font(bold=True, color='FFFFFF', size=10)
    data_font = Font(size=9)
    headers = ["No", "지역", "사건번호", "물건번호", "소재지", "면적", "용도",
               "감정평가액", "최저매각가격", "매각가율", "매각기일", "담당계", "진행상태", "비고"]
    widths = [4, 5, 18, 6, 55, 12, 18, 15, 15, 8, 12, 8, 10, 15]

    def write_sheet(ws, data, fill_color):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor=fill_color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for idx, item in enumerate(data, 1):
            row = idx + 1
            region = get_region(item.get('address', ''))
            try:
                appr = int(item.get('appraisal', '0').replace(',', ''))
            except (ValueError, AttributeError):
                appr = 0
            try:
                minp = int(item.get('minPrice', '0').replace(',', ''))
            except (ValueError, AttributeError):
                minp = 0
            ratio = item.get('ratio', '')
            vals = [idx, region, item.get('caseNo',''), item.get('itemNo',''),
                    item.get('address',''), item.get('detail',''), item.get('usage',''),
                    appr, minp, ratio, item.get('saleDate',''),
                    item.get('dept',''), item.get('status',''), item.get('note','')]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = data_font
                cell.border = border
                if col in (8, 9):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                elif col in (1, 2, 4, 10, 11, 12, 13):
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(wrap_text=True)
        ws.auto_filter.ref = f"A1:N{len(data)+1}"
        ws.freeze_panes = 'A2'

    ws1 = wb.active
    ws1.title = f"전체 매물 ({len(items)}건)"[:31]
    write_sheet(ws1, items, '2F5496')

    filtered = []
    if region_filter or usage_filter:
        filtered = items
        if region_filter:
            filtered = [i for i in filtered if region_filter in i.get('address', '')]
        if usage_filter:
            filtered = [i for i in filtered if usage_filter in i.get('usage', '')]
        label = ' '.join(filter(None, [region_filter, usage_filter]))
        sheet_name = f"{label} ({len(filtered)}건)"[:31]
        ws2 = wb.create_sheet(sheet_name)
        write_sheet(ws2, filtered, '548235')

    wb.save(output_path)
    print(f"Created: {output_path}")
    print(f"Total: {len(items)} items")
    if filtered:
        print(f"Filtered: {len(filtered)} items")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('input_json')
    parser.add_argument('output_xlsx')
    parser.add_argument('--region', default=None)
    parser.add_argument('--usage', default=None)
    args = parser.parse_args()
    with open(args.input_json) as f:
        raw = json.load(f)
    items = normalize_input(raw)
    create_excel(items, args.output_xlsx, args.region, args.usage)
